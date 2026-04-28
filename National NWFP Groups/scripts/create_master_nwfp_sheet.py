from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


WORKBOOK_PATH = Path(__file__).resolve().parent.parent / "NWFP Group-National Data.xlsx"
SOURCE_SHEET = "Dzongkhag Wise-NWFP Groups"
TARGET_SHEET = "Master-NWFP"

MASTER_HEADERS = [
    "SN",
    "Group Name",
    "Dzongkhag",
    "Division/Park",
    "Gewog",
    "Village",
    "Members (Nos)",
    "Female (Nos)",
    "Area (ha)",
    "Estd year",
    "Revision Year",
    "Plan Period",
    "Plan Type",
    "Contact Details",
    "Species",
    "Review Date",
    "Review Status",
    "Source Sheet",
    "Source Row",
    "Legacy Female (nos)",
    "Legacy Area (ha)",
    "Legacy Estd year",
    "Legacy Revision",
    "Legacy Plan Year",
    "Legacy Contact Details",
    "Legacy Species",
    "Data Notes",
]

LEGACY_HEADERS = MASTER_HEADERS[19:26]
CONTACT_KEYWORDS = (
    "chairperson",
    "secretary",
    "treasurer",
    "harvested",
    "contact",
    "mobile",
    "phone",
)


def clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def to_text(value) -> str:
    value = clean(value)
    if value is None:
        return ""
    return str(value).strip()


def looks_plan_period(value) -> bool:
    text = to_text(value)
    if not text:
        return False
    lower = text.lower()
    if "plan expired" in lower:
        return True
    if re.search(r"\b(19|20)\d{2}\s*[-/]\s*(19|20)\d{2}\b", text):
        return True
    if re.search(r"\d{1,2}/\d{1,2}/(19|20)\d{2}", text):
        return True
    return False


def looks_year_note(value) -> bool:
    text = to_text(value)
    if not text or looks_plan_period(text):
        return False
    if re.fullmatch(r"\d{4}", text):
        year = int(text)
        return 1900 <= year <= 2100
    if re.search(r"\b(19|20)\d{2}\b", text):
        return True
    return False


def looks_contact(value) -> bool:
    text = to_text(value)
    if not text:
        return False
    lower = text.lower()
    if any(keyword in lower for keyword in CONTACT_KEYWORDS):
        return True
    return bool(re.search(r"\d{6,}", text))


def looks_area(value) -> bool:
    value = clean(value)
    if value is None:
        return False
    text = to_text(value).replace(",", ".")
    if not re.fullmatch(r"\d+(?:\.\d+)?", text):
        return False
    number = float(text)
    return not number.is_integer() or not (1900 <= int(number) <= 2100)


def join_parts(parts: list[str]) -> str | None:
    cleaned = [part for part in (to_text(value) for value in parts) if part]
    if not cleaned:
        return None
    return " | ".join(cleaned)


def build_master_row(raw_row: tuple, source_row_number: int) -> list:
    legacy_values = [clean(value) for value in raw_row[5:12]]
    notes: list[str] = []
    area = None
    estd_years: list = []
    plan_period = None
    contact = None
    species_parts: list[str] = []
    area_source = None

    for offset, value in enumerate(legacy_values, start=6):
        if clean(value) is None:
            continue
        if looks_contact(value):
            if contact is None:
                contact = clean(value)
            else:
                species_parts.append(to_text(value))
            continue
        if looks_plan_period(value):
            if plan_period is None:
                plan_period = clean(value)
            else:
                species_parts.append(to_text(value))
            continue
        if looks_area(value):
            if area is None:
                area = clean(value)
                area_source = offset
            else:
                notes.append(f"Extra numeric area-like value kept in legacy column {offset}.")
            continue
        if looks_year_note(value):
            estd_years.append(clean(value))
            continue
        species_parts.append(to_text(value))

    if area_source == 6:
        notes.append("Area parsed from legacy 'Female (nos)' column.")
    elif area_source == 7:
        notes.append("Area parsed from legacy 'Area (ha)' column.")

    if raw_row[5] is not None:
        notes.append("Female (Nos) left blank because sheet 1 legacy column F is not reliably female-count data.")

    if species_parts:
        if legacy_values[4] and to_text(legacy_values[4]) in species_parts:
            notes.append("Species parsed from legacy 'Plan Year' column.")
        if legacy_values[6] and to_text(legacy_values[6]) in species_parts:
            notes.append("Species parsed from legacy 'Species' column.")

    return [
        clean(raw_row[0]),
        clean(raw_row[1]),
        clean(raw_row[2]),
        None,
        clean(raw_row[3]),
        None,
        clean(raw_row[4]),
        None,
        area,
        estd_years[0] if estd_years else None,
        estd_years[1] if len(estd_years) > 1 else None,
        plan_period,
        None,
        contact,
        join_parts(species_parts),
        None,
        None,
        SOURCE_SHEET,
        source_row_number,
        *legacy_values,
        join_parts(notes),
    ]


def format_sheet(worksheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    wrap_alignment = Alignment(vertical="top", wrap_text=True)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap_alignment

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    widths = {
        "A": 8,
        "B": 42,
        "C": 18,
        "D": 18,
        "E": 18,
        "F": 22,
        "G": 16,
        "H": 14,
        "I": 14,
        "J": 16,
        "K": 16,
        "L": 18,
        "M": 12,
        "N": 28,
        "O": 60,
        "P": 14,
        "Q": 16,
        "R": 28,
        "S": 12,
        "T": 16,
        "U": 16,
        "V": 16,
        "W": 16,
        "X": 18,
        "Y": 28,
        "Z": 28,
        "AA": 60,
    }

    for column_letter, width in widths.items():
        worksheet.column_dimensions[column_letter].width = width

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap_alignment


def create_master_sheet() -> int:
    workbook = load_workbook(WORKBOOK_PATH)
    source = workbook[SOURCE_SHEET]

    if TARGET_SHEET in workbook.sheetnames:
        workbook.remove(workbook[TARGET_SHEET])

    target = workbook.create_sheet(TARGET_SHEET)
    target.append(MASTER_HEADERS)

    written_rows = 0
    for row_number, row in enumerate(
        source.iter_rows(min_row=2, max_col=12, values_only=True),
        start=2,
    ):
        group_name = to_text(row[1])
        if not group_name or group_name.upper() == "INACTIVE":
            continue
        target.append(build_master_row(row, row_number))
        written_rows += 1

    format_sheet(target)
    workbook.save(WORKBOOK_PATH)
    return written_rows


if __name__ == "__main__":
    created = create_master_sheet()
    print(f"Created '{TARGET_SHEET}' with {created} rows in {WORKBOOK_PATH.name}.")
